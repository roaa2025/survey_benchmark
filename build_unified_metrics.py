#!/usr/bin/env python3
"""
Unified metrics pipeline: consolidates GPT-4o telemetry (TXT) and DeepSeek evaluation data (Excel)
into a single JSON schema for model comparison.
"""

import argparse
import hashlib
import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    exit(1)


def extract_duration_ms(entry: Dict[str, Any]) -> Optional[int]:
    """
    Extract duration in milliseconds using comprehensive fallback chain:
    1. duration_ms
    2. duration_sec * 1000
    3. duration (if < 100, assume seconds; else ms)
    4. timing.duration_ms (nested)
    5. compute from start_time/end_time if present
    6. Check for common variations: elapsed_time, execution_time, response_time, latency
    """
    # Try duration_ms directly
    duration_ms = entry.get('duration_ms')
    if duration_ms is not None:
        try:
            return int(float(duration_ms))
        except (ValueError, TypeError):
            pass
    
    # Try duration_sec * 1000
    duration_sec = entry.get('duration_sec')
    if duration_sec is not None:
        try:
            return int(float(duration_sec) * 1000)
        except (ValueError, TypeError):
            pass
    
    # Try duration (if < 100, assume seconds; else ms)
    duration = entry.get('duration')
    if duration is not None:
        try:
            duration_val = float(duration)
            if duration_val < 100:
                return int(duration_val * 1000)
            else:
                return int(duration_val)
        except (ValueError, TypeError):
            pass
    
    # Try timing.duration_ms (nested)
    timing = entry.get('timing')
    if isinstance(timing, dict):
        timing_duration = timing.get('duration_ms')
        if timing_duration is not None:
            try:
                return int(float(timing_duration))
            except (ValueError, TypeError):
                pass
    
    # Try computing from start_time/end_time
    start_time = entry.get('start_time')
    end_time = entry.get('end_time')
    if start_time is not None and end_time is not None:
        try:
            start = float(start_time)
            end = float(end_time)
            duration_ms = int((end - start) * 1000)
            if duration_ms > 0:
                return duration_ms
        except (ValueError, TypeError):
            pass
    
    # Try alternative field names
    for field_name in ['elapsed_time', 'execution_time', 'response_time', 'latency', 'time_ms', 'time_sec']:
        field_val = entry.get(field_name)
        if field_val is not None:
            try:
                val = float(field_val)
                if 'sec' in field_name or val < 100:
                    return int(val * 1000)
                else:
                    return int(val)
            except (ValueError, TypeError):
                continue
    
    return None


def extract_tokens(entry: Dict[str, Any]) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    """
    Extract token usage using comprehensive fallback chain:
    For total_tokens: token_usage.total_tokens -> total_tokens -> usage.total_tokens
    For input_tokens: token_usage.prompt_tokens -> token_usage.input_tokens -> input_tokens -> usage.input_tokens
    For output_tokens: token_usage.completion_tokens -> token_usage.output_tokens -> output_tokens -> usage.output_tokens
    Also checks for alternative field names and nested structures.
    Returns: (tokens_total, tokens_input, tokens_output)
    """
    # Extract total_tokens
    tokens_total = None
    token_usage = entry.get('token_usage')
    if isinstance(token_usage, dict):
        tokens_total = token_usage.get('total_tokens')
    
    if tokens_total is None:
        tokens_total = entry.get('total_tokens')
    
    if tokens_total is None:
        usage = entry.get('usage')
        if isinstance(usage, dict):
            tokens_total = usage.get('total_tokens')
    
    # Extract input tokens
    tokens_input = None
    if isinstance(token_usage, dict):
        tokens_input = token_usage.get('prompt_tokens')
        if tokens_input is None:
            tokens_input = token_usage.get('input_tokens')
    
    if tokens_input is None:
        tokens_input = entry.get('input_tokens')
    
    if tokens_input is None:
        usage = entry.get('usage')
        if isinstance(usage, dict):
            tokens_input = usage.get('input_tokens')
    
    # Extract output tokens
    tokens_output = None
    if isinstance(token_usage, dict):
        tokens_output = token_usage.get('completion_tokens')
        if tokens_output is None:
            tokens_output = token_usage.get('output_tokens')
    
    if tokens_output is None:
        tokens_output = entry.get('output_tokens')
    
    if tokens_output is None:
        usage = entry.get('usage')
        if isinstance(usage, dict):
            tokens_output = usage.get('output_tokens')
    
    # Try alternative field names for tokens
    if tokens_total is None:
        for alt_name in ['tokens', 'num_tokens', 'token_count', 'total_token_count']:
            alt_val = entry.get(alt_name)
            if alt_val is not None:
                try:
                    tokens_total = int(float(alt_val))
                    break
                except (ValueError, TypeError):
                    continue
    
    # Convert to int if not None
    try:
        tokens_total = int(tokens_total) if tokens_total is not None else None
    except (ValueError, TypeError):
        tokens_total = None
    
    try:
        tokens_input = int(tokens_input) if tokens_input is not None else None
    except (ValueError, TypeError):
        tokens_input = None
    
    try:
        tokens_output = int(tokens_output) if tokens_output is not None else None
    except (ValueError, TypeError):
        tokens_output = None
    
    return tokens_total, tokens_input, tokens_output


def parse_gpt_txt(file_path: str) -> List[Dict[str, Any]]:
    """
    Parse GPT-4o telemetry TXT file and extract EVAL_METRICS events.
    Filters for generation runs (flow like 'fast_generate') and extracts relevant fields.
    """
    if not os.path.exists(file_path):
        print(f"Warning: GPT TXT file not found: {file_path}")
        return []
    
    with open(file_path, 'r', encoding='utf-8') as f:
        text = f.read()
    
    json_blocks = []
    json_regex = re.compile(r'Full JSON:\s*(\{[\s\S]*?\})', re.MULTILINE)
    
    for match in json_regex.finditer(text):
        try:
            json_str = match.group(1)
            parsed = json.loads(json_str)
            if parsed.get('event') == 'EVAL_METRICS':
                json_blocks.append(parsed)
        except (json.JSONDecodeError, KeyError) as e:
            # Suppress verbose JSON parsing warnings - they're expected for malformed blocks
            continue
    
    runs = []
    sample_entry_keys = None
    for idx, entry in enumerate(json_blocks):
        flow = entry.get('flow', '')
        
        # Filter for generation runs (flow contains 'generate')
        if 'generate' not in flow.lower():
            continue
        
        # Capture sample entry keys for debugging (first entry only)
        if sample_entry_keys is None and idx == 0:
            sample_entry_keys = list(entry.keys())
            print(f"Sample EVAL_METRICS entry keys: {sample_entry_keys}")
        
        thread_id = entry.get('thread_id')
        
        # Extract duration using comprehensive fallback chain
        duration_ms = extract_duration_ms(entry)
        
        # Extract token usage using comprehensive fallback chain
        tokens_total, tokens_input, tokens_output = extract_tokens(entry)
        
        # Extract output shape
        questions = entry.get('questions_total')
        pages = entry.get('pages_total')
        rules = entry.get('rules_total')
        
        # Extract invalid question types as list
        invalid_question_types = []
        invalid_breakdown = entry.get('invalid_question_type_breakdown', {})
        if isinstance(invalid_breakdown, dict):
            invalid_question_types = list(invalid_breakdown.keys())
        
        # Determine status
        has_errors = (
            entry.get('schema_error_count', 0) > 0 or
            entry.get('missing_required_fields_count', 0) > 0 or
            entry.get('rules_invalid_ref_count', 0) > 0 or
            entry.get('rules_schema_error_count', 0) > 0
        )
        status = "fail" if has_errors else "success"
        
        # Generate stable run_id
        run_id_parts = [
            "gpt4o_eval_metrics_all_txt",
            str(thread_id) if thread_id else "",
            str(duration_ms) if duration_ms else "",
            str(tokens_total) if tokens_total else "",
            str(questions) if questions else "",
            str(pages) if pages else "",
            str(idx)
        ]
        run_id = hashlib.sha256("|".join(run_id_parts).encode()).hexdigest()[:16]
        
        # Build run object
        run = {
            "run_id": run_id,
            "source_id": "gpt4o_eval_metrics_all_txt",
            "model": {
                "family": "gpt",
                "name": "gpt-4o",
                "variant": None,
                "role": None
            },
            "task": {
                "suite": "generation",
                "scenario_id": None,
                "thread_id": str(thread_id) if thread_id else None,
                "language": None,
                "prompt": {
                    "text": None,
                    "length_chars": None
                }
            },
            "timing": {
                "duration_ms": int(duration_ms) if duration_ms is not None else None
            },
            "usage": {
                "tokens_total": int(tokens_total) if tokens_total is not None else None,
                "tokens_input": int(tokens_input) if tokens_input is not None else None,
                "tokens_output": int(tokens_output) if tokens_output is not None else None
            },
            "output_shape": {
                "pages": int(pages) if pages is not None else None,
                "questions": int(questions) if questions is not None else None,
                "rules": int(rules) if rules is not None else None,
                "invalid_question_types": invalid_question_types
            },
            "quality": {
                "llm_judge": {
                    "overall": None,
                    "question_quality": None,
                    "survey_coherence": None,
                    "bilingual_alignment": None,
                    "question_page_distribution": None,
                    "controller_appropriateness": None
                }
            },
            "stability": {
                "status": status,
                "telemetry": {
                    "has_duration": duration_ms is not None,
                    "has_tokens": tokens_total is not None,
                    "has_judge_scores": False
                }
            }
        }
        
        runs.append(run)
    
    return runs


def normalize_language(lang: Optional[str]) -> Optional[str]:
    """Normalize language value to ar/en/bilingual or null."""
    if not lang:
        return None
    
    lang_lower = str(lang).lower().strip()
    if lang_lower in ['ar', 'arabic']:
        return 'ar'
    elif lang_lower in ['en', 'english']:
        return 'en'
    elif lang_lower in ['bilingual', 'bi', 'both']:
        return 'bilingual'
    else:
        return None


def parse_deepseek_excel(file_path: str) -> List[Dict[str, Any]]:
    """
    Parse DeepSeek evaluation Excel file and extract run data.
    Maps columns to unified schema fields.
    """
    if not os.path.exists(file_path):
        print(f"Warning: DeepSeek Excel file not found: {file_path}")
        return []
    
    try:
        wb = load_workbook(file_path, data_only=True)
        # Use first sheet
        ws = wb.active
        
        # Find header row - try first row by default, or search for column names
        headers = {}
        header_row = None
        
        # First, try the first row
        first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        first_row_values = [str(cell).lower().strip() if cell else "" for cell in first_row]
        
        # Check if first row looks like headers (contains common header keywords)
        if any(keyword in ' '.join(first_row_values) for keyword in ['scenario', 'language', 'question', 'page', 'score', 'time']):
            header_row = 1
            for col_idx, cell in enumerate(first_row, 1):
                if cell:
                    header_key = str(cell).lower().strip()
                    headers[header_key] = col_idx
        else:
            # Search for header row in first 10 rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                row_values = [str(cell).lower().strip() if cell else "" for cell in row]
                if any(keyword in ' '.join(row_values) for keyword in ['scenario', 'language', 'question', 'page', 'score']):
                    header_row = row_idx
                    for col_idx, cell in enumerate(row, 1):
                        if cell:
                            header_key = str(cell).lower().strip()
                            headers[header_key] = col_idx
                    break
        
        if not header_row or not headers:
            print(f"Warning: Could not find header row in Excel file. First row values: {first_row_values[:5]}")
            print("Trying first row as headers anyway...")
            header_row = 1
            for col_idx, cell in enumerate(first_row, 1):
                if cell:
                    header_key = str(cell).lower().strip()
                    headers[header_key] = col_idx
            if not headers:
                return []
        
        runs = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            # Skip empty rows
            if not any(cell for cell in row):
                continue
            
            # Extract values by column name (case-insensitive matching with whitespace handling)
            def get_col(keywords: List[str], default=None):
                for keyword in keywords:
                    keyword_lower = keyword.lower().strip()
                    for header_key, col_idx in headers.items():
                        # Match if keyword is contained in header (substring match)
                        if keyword_lower in header_key:
                            val = row[col_idx - 1] if col_idx <= len(row) else None
                            # Handle empty strings, None, and whitespace-only values
                            if val is None:
                                return default
                            val_str = str(val).strip()
                            if val_str == '' or val_str.lower() == 'none' or val_str.lower() == 'n/a':
                                return default
                            return val
                return default
            
            scenario_id = get_col(['scenario id', 'scenario_id', 'scenario'])
            language = normalize_language(get_col(['survey language', 'language', 'lang']))
            user_prompt = get_col(['user prompt', 'prompt', 'user_prompt'])
            prompt_length = get_col(['prompt length', 'prompt_length', 'length'])
            generation_time = get_col(['generation time', 'generation_time', 'time', 'duration'])
            # Extract questions and pages - prioritize exact matches for "Number of Questions" and "Number of Pages"
            num_questions = get_col(['number of questions', 'questions', 'num_questions', 'question_count', 'question'])
            num_pages = get_col(['number of pages', 'pages', 'num_pages', 'page_count', 'page'])
            overall_score = get_col(['overall_score', 'overall score', 'overall'])
            question_quality = get_col(['question quality score', 'question_quality', 'question quality'])
            survey_coherence = get_col(['survey coherence score', 'survey_coherence', 'survey coherence'])
            bilingual_alignment = get_col(['bilingual alignment score', 'bilingual_alignment', 'bilingual alignment'])
            question_page_dist = get_col(['question page distribution score', 'question_page_distribution', 'question page distribution'])
            controller_appropriateness = get_col(['controller appropriateness score', 'controller_appropriateness', 'controller appropriateness'])
            
            # Extract token usage - try multiple column name variations
            tokens_total = get_col(['tokens_total', 'total_tokens', 'tokens', 'token count', 'total token count'])
            tokens_input = get_col(['tokens_input', 'input_tokens', 'prompt_tokens', 'prompt tokens', 'input token count'])
            tokens_output = get_col(['tokens_output', 'output_tokens', 'completion_tokens', 'completion tokens', 'output token count'])
            
            # Convert generation_time to ms (if < 1000, assume seconds)
            duration_ms = None
            if generation_time is not None:
                try:
                    gen_time = float(generation_time)
                    if gen_time < 1000:
                        duration_ms = int(gen_time * 1000)
                    else:
                        duration_ms = int(gen_time)
                except (ValueError, TypeError):
                    pass
            
            # Get prompt length
            prompt_length_chars = None
            if prompt_length is not None:
                try:
                    prompt_length_chars = int(prompt_length)
                except (ValueError, TypeError):
                    if user_prompt:
                        prompt_length_chars = len(str(user_prompt))
            elif user_prompt:
                prompt_length_chars = len(str(user_prompt))
            
            # Convert numeric fields
            def to_int_or_none(val):
                try:
                    return int(float(val)) if val is not None else None
                except (ValueError, TypeError):
                    return None
            
            def to_float_or_none(val):
                try:
                    return float(val) if val is not None else None
                except (ValueError, TypeError):
                    return None
            
            # Generate run_id
            run_id_parts = [
                "deepseek_eval_excel",
                str(scenario_id) if scenario_id else "",
                str(duration_ms) if duration_ms else "",
                str(num_questions) if num_questions else "",
                str(num_pages) if num_pages else "",
                str(row_idx)
            ]
            run_id = hashlib.sha256("|".join(run_id_parts).encode()).hexdigest()[:16]
            
            # Build run object
            run = {
                "run_id": run_id,
                "source_id": "deepseek_eval_excel",
                "model": {
                    "family": "deepseek",
                    "name": "deepseek",
                    "variant": None,
                    "role": None
                },
                "task": {
                    "suite": "generation",
                    "scenario_id": str(scenario_id) if scenario_id else None,
                    "thread_id": None,
                    "language": language,
                    "prompt": {
                        "text": str(user_prompt) if user_prompt else None,
                        "length_chars": prompt_length_chars
                    }
                },
                "timing": {
                    "duration_ms": duration_ms
                },
                "usage": {
                    "tokens_total": to_int_or_none(tokens_total),
                    "tokens_input": to_int_or_none(tokens_input),
                    "tokens_output": to_int_or_none(tokens_output)
                },
                "output_shape": {
                    "pages": to_int_or_none(num_pages),
                    "questions": to_int_or_none(num_questions),
                    "rules": None,
                    "invalid_question_types": []
                },
                "quality": {
                    "llm_judge": {
                        "overall": to_float_or_none(overall_score),
                        "question_quality": to_float_or_none(question_quality),
                        "survey_coherence": to_float_or_none(survey_coherence),
                        "bilingual_alignment": to_float_or_none(bilingual_alignment),
                        "question_page_distribution": to_float_or_none(question_page_dist),
                        "controller_appropriateness": to_float_or_none(controller_appropriateness)
                    }
                },
                "stability": {
                    "status": "unknown",
                    "telemetry": {
                        "has_duration": duration_ms is not None,
                        "has_tokens": tokens_total is not None,
                        "has_judge_scores": overall_score is not None
                    }
                }
            }
            
            runs.append(run)
        
        wb.close()
        return runs
    
    except Exception as e:
        print(f"Error parsing Excel file: {e}")
        return []


def build_unified_metrics(gpt_runs: List[Dict], deepseek_runs: List[Dict]) -> Dict[str, Any]:
    """Assemble unified JSON with schema_version, generated_at, sources, and runs."""
    return {
        "schema_version": "1.0.0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "sources": [
            {
                "id": "gpt4o_eval_metrics_all_txt",
                "type": "telemetry_txt",
                "runs_count": len(gpt_runs)
            },
            {
                "id": "deepseek_eval_excel",
                "type": "evaluation_excel",
                "runs_count": len(deepseek_runs)
            }
        ],
        "runs": gpt_runs + deepseek_runs
    }


def compute_summary_stats(runs: List[Dict]) -> Dict[str, Any]:
    """Calculate coverage percentages for tokens, duration, and judge scores."""
    total = len(runs)
    if total == 0:
        return {
            "token_coverage": 0.0,
            "duration_coverage": 0.0,
            "judge_coverage": 0.0
        }
    
    has_tokens = sum(1 for r in runs if r.get("usage", {}).get("tokens_total") is not None)
    has_duration = sum(1 for r in runs if r.get("timing", {}).get("duration_ms") is not None)
    has_judge = sum(1 for r in runs if r.get("quality", {}).get("llm_judge", {}).get("overall") is not None)
    
    return {
        "token_coverage": (has_tokens / total * 100) if total > 0 else 0.0,
        "duration_coverage": (has_duration / total * 100) if total > 0 else 0.0,
        "judge_coverage": (has_judge / total * 100) if total > 0 else 0.0
    }


def find_input_files() -> Tuple[Optional[str], Optional[str]]:
    """Try to find input files in common locations."""
    root = Path(__file__).parent
    
    # Try common data directory locations
    possible_data_dirs = [
        Path("C:/Users/roaa.alashqar/Desktop/eval draft data"),
        root / "data",
        root.parent / "eval draft data"
    ]
    
    gpt_txt = None
    deepseek_xlsx = None
    
    # Check root directory first
    for path in [root / "EVAL_METRICS_All.txt", root / "data" / "EVAL_METRICS_All.txt"]:
        if path.exists():
            gpt_txt = str(path)
            break
    
    # Check data directories
    if not gpt_txt:
        for data_dir in possible_data_dirs:
            if data_dir.exists():
                txt_path = data_dir / "EVAL_METRICS_All.txt"
                if txt_path.exists():
                    gpt_txt = str(txt_path)
                    break
    
    # Check root directory for Excel
    for path in [root / "Test Scenarios and Eval Dataset.xlsx", root / "data" / "Test Scenarios and Eval Dataset.xlsx"]:
        if path.exists():
            deepseek_xlsx = str(path)
            break
    
    # Check data directories for Excel
    if not deepseek_xlsx:
        for data_dir in possible_data_dirs:
            if data_dir.exists():
                xlsx_path = data_dir / "Test Scenarios and Eval Dataset.xlsx"
                if xlsx_path.exists():
                    deepseek_xlsx = str(xlsx_path)
                    break
    
    return gpt_txt, deepseek_xlsx


def main():
    parser = argparse.ArgumentParser(description="Build unified metrics JSON from GPT TXT and DeepSeek Excel")
    parser.add_argument("--gpt_txt", type=str, help="Path to EVAL_METRICS_All.txt")
    parser.add_argument("--deepseek_xlsx", type=str, help="Path to Test Scenarios and Eval Dataset.xlsx")
    parser.add_argument("--out", type=str, default="reports/metrics_unified.json", help="Output JSON file path")
    
    args = parser.parse_args()
    
    # Find input files if not provided
    gpt_txt = args.gpt_txt
    deepseek_xlsx = args.deepseek_xlsx
    
    if not gpt_txt or not deepseek_xlsx:
        found_gpt, found_deepseek = find_input_files()
        if not gpt_txt:
            gpt_txt = found_gpt
        if not deepseek_xlsx:
            deepseek_xlsx = found_deepseek
    
    # Parse files
    print("Parsing GPT TXT file...")
    gpt_runs = parse_gpt_txt(gpt_txt) if gpt_txt else []
    print(f"  Extracted {len(gpt_runs)} GPT runs")
    
    print("Parsing DeepSeek Excel file...")
    deepseek_runs = parse_deepseek_excel(deepseek_xlsx) if deepseek_xlsx else []
    print(f"  Extracted {len(deepseek_runs)} DeepSeek runs")
    
    # Build unified metrics
    unified = build_unified_metrics(gpt_runs, deepseek_runs)
    
    # Compute summary stats
    all_runs = gpt_runs + deepseek_runs
    stats = compute_summary_stats(all_runs)
    
    # Write output
    output_path = Path(args.out)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(unified, f, indent=2, ensure_ascii=False)
    
    # Print summary
    print(f"\nSuccessfully generated {output_path}")
    print(f"Total runs: {len(all_runs)}")
    print(f"  GPT-4o: {len(gpt_runs)}")
    print(f"  DeepSeek: {len(deepseek_runs)}")
    print(f"\nCoverage:")
    print(f"  Token coverage: {stats['token_coverage']:.1f}%")
    print(f"  Duration coverage: {stats['duration_coverage']:.1f}%")
    print(f"  Judge coverage: {stats['judge_coverage']:.1f}%")


if __name__ == "__main__":
    main()

